import json
import datetime
import pandas as pd

from collections import defaultdict

from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.utils import timezone

from django.db.models import Sum, Max
from django.db.models.functions import ExtractWeek, ExtractIsoYear

from openpyxl import Workbook
from .auxiliares import *
from plantaE.models import (
    AcumFruta,
    Recepciones,
    Boletas,
    datosProduccion,
    usuariosAppFruta,
    detallesEstructuras,
    detallerec,
    detallerecaux
)

def exportar_excel(request):
    if request.method == 'POST':
        opcion1 = request.POST.get('opcion1')
        # Crea un libro de Excel y una hoja
        wb = Workbook()
        ws = wb.active
        ws.title = 'Rio'

        # Obtén los datos de tu modelo
        datos_rio = AcumFruta.objects.filter(fecha=opcion1, finca="RIO").exclude(status='Anulado').values(
            "id","fecha", "finca", "orden", "cultivo", "variedad", "estructura", "cajas","correo","libras"
        ).order_by("orden")

        # Agrega los encabezados
        ws.append(["id","fecha", "finca", "orden", "cultivo", "variedad", "estructura", "cajas","correo","libras"])
        
        # Agrega los datos  
        for obj in datos_rio:
            row = [obj['id'],obj['fecha'], obj['finca'], obj['orden'], obj['cultivo'], obj['variedad'], obj['estructura'], obj['cajas'], obj['correo'], obj['libras']]
            ws.append(row)

        ws_valle = wb.create_sheet(title='Valle')
        
        # Filtra tus datos según la opción seleccionada
        datos_valle = AcumFruta.objects.filter(fecha=opcion1, finca="VALLE").exclude('Anulado').values(
            "id", "fecha", "finca", "orden", "cultivo", "variedad", "estructura", "cajas","correo","libras"
        )

        # Crea un DataFrame a partir de los datos
        df = pd.DataFrame(list(datos_valle))
        if not df.empty:
            # Agrupa los datos
            df_agrupado = df.groupby(['orden', 'estructura', 'variedad'], as_index=False).agg(
                id=('id', 'first'),
                fecha=('fecha', 'first'),
                finca=('finca', 'first'),
                orden=('orden', 'first'),
                cultivo=('cultivo', 'first'),
                variedad=('variedad', 'first'),
                estructura=('estructura', 'first'),
                total_cajas=('cajas', 'sum'),
                total_libras=('libras', 'sum'),
                correo = ('correo', 'first'),
            )

            df_agrupado = df_agrupado.sort_values(by='orden')

            # Agrega encabezados a la hoja Valle
            ws_valle.append(df_agrupado.columns.tolist())

            # Agrega los registros agrupados a la hoja Valle
            for record in df_agrupado.itertuples(index=False):
                ws_valle.append(record)  # Excluir el índice, si es necesario

        
        ws_provalle = wb.create_sheet(title='Provalle')
        
        # Filtra tus datos según la opción seleccionada
        datos_provalle = AcumFruta.objects.filter(fecha=opcion1, finca="PRODUCTOS DEL VALLE, S.A.").exclude(status='Anulado').values(
            "id", "fecha", "finca", "orden", "cultivo", "variedad", "estructura", "cajas","correo","libras"
        )
        df = pd.DataFrame(list(datos_provalle))
        if not df.empty:
            # Crea un DataFrame a partir de los datos

            # Agrupa los datos
            df_agrupado = df.groupby(['orden', 'estructura', 'variedad'], as_index=False).agg(
                id=('id', 'first'),
                fecha=('fecha', 'first'),
                finca=('finca', 'first'),
                orden=('orden', 'first'),
                cultivo=('cultivo', 'first'),
                variedad=('variedad', 'first'),
                estructura=('estructura', 'first'),
                total_cajas=('cajas', 'sum'),
                total_libras=('libras', 'sum'),
                correo = ('correo', 'first'),
            )

            df_agrupado = df_agrupado.sort_values(by='orden')

            # Agrega encabezados a la hoja Valle
            ws_provalle.append(df_agrupado.columns.tolist())

            # Agrega los registros agrupados a la hoja Valle
            for record in df_agrupado.itertuples(index=False):
                ws_provalle.append(record)  # Excluir el índice, si es necesario
        
        # Crea una respuesta HTTP que sirva el archivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=envios.xlsx'

        # Guarda el libro de Excel en la respuesta
        wb.save(response)

        return response

    return render(request, 'plantaE/reportetecnicos/consulta_envios.html')

def dashboard_tecnicos(request):
    # Filtros desde GET
    filtros_get = {
        'finca': request.GET.get('finca'),
        'orden': request.GET.get('orden'),
        'variedad': request.GET.get('variedad'),
        'cultivo': request.GET.get('cultivo'),
        'estructura': request.GET.get('estructura'),
    }
    nombre_usuario = request.user.username
    ordenes_abiertas = datosProduccion.objects.filter(status='Abierta').values_list('orden', flat=True)

    # Query base
    qs = AcumFruta.objects.filter(correo=nombre_usuario,orden__in=ordenes_abiertas).exclude(finca="CIP").exclude(libras__isnull=True)

    # Aplicar filtros
    for campo, valor in filtros_get.items():
        if valor:
            qs = qs.filter(**{campo: valor})

    # Agrupación por semana
    datos = qs.annotate(
        semana=ExtractWeek('fecha'),
        anio=ExtractIsoYear('fecha')
    ).values('anio', 'semana').annotate(
        libras_totales=Sum('libras')
    ).order_by('anio', 'semana')

    # Ejes para la gráfica
    fechas = [get_date_from_week(d['anio'], d['semana']).isoformat() for d in datos]
    kilos = [round(d['libras_totales'] / 2.20462, 2) for d in datos]
    derivadas = [0] + [kilos[i] - kilos[i - 1] for i in range(1, len(kilos))]


    # Filtros disponibles
    filtros_completos = [
        ('Finca', 'finca', AcumFruta.objects.filter(correo=nombre_usuario,orden__in=ordenes_abiertas).exclude(finca__isnull=True).exclude(finca='').values_list('finca', flat=True).distinct()),
        ('Orden', 'orden', AcumFruta.objects.filter(correo=nombre_usuario,orden__in=ordenes_abiertas).exclude(orden__isnull=True).exclude(orden='').values_list('orden', flat=True).distinct()),
        ('Variedad', 'variedad', AcumFruta.objects.filter(correo=nombre_usuario,orden__in=ordenes_abiertas).exclude(variedad__isnull=True).exclude(variedad='').values_list('variedad', flat=True).distinct()),
        ('Cultivo', 'cultivo', AcumFruta.objects.filter(correo=nombre_usuario,orden__in=ordenes_abiertas).exclude(cultivo__isnull=True).exclude(cultivo='').values_list('cultivo', flat=True).distinct()),
        ('Estructura', 'estructura', AcumFruta.objects.filter(correo=nombre_usuario,orden__in=ordenes_abiertas).exclude(estructura__isnull=True).exclude(estructura='').values_list('estructura', flat=True).distinct()),
    ]

    context = {
        'filtros_completos': filtros_completos,
        'fechas_json': json.dumps(fechas),
        'libras_json': json.dumps(kilos),
        'derivadas_json': json.dumps(derivadas),
        'request': request,
    }

    return render(request, 'plantaE/reportetecnicos/dashboard_acumfruta2.html', context)

def semanalprodterm_pivot_productor(request):

    nombre_usuario = request.user.username
    datos = usuariosAppFruta.objects.filter(correo=nombre_usuario).values('finca', 'encargado')

    # Definir fecha límite: 31 de octubre de 2025
    fecha_limite = datetime.date(2025, 9, 30)

    finca_usuario = datos[0]['finca']
    inventario_datos = Boletas.objects.filter(finca=finca_usuario,fecha__gt=fecha_limite).annotate(
        semana=ExtractWeek('fecha'),
        anio=ExtractIsoYear('fecha')
    ).values('itemsapname', 'categoria', 'cultivo', 'semana', 'anio').annotate(
        total_libras=Sum('libras'),total_cajas=Sum('cajas')
    ).order_by('anio', 'semana', 'itemsapname', 'categoria')

    resultado = []
    total_por_semana_cultivo = defaultdict(float)

    for registro in inventario_datos:
        clave = (registro['anio'], registro['semana'], registro['cultivo'])
        total_libras = registro['total_libras'] or 0
        total_cajas = registro['total_cajas'] or 0
        total_por_semana_cultivo[clave] += total_libras
        kilos = total_libras / 2.20462 if total_libras else 0

    for registro in inventario_datos:
        clave = (registro['anio'], registro['semana'], registro['cultivo'])
        total_cultivo_semana = total_por_semana_cultivo[clave]
        total_cajas = registro['total_cajas'] or 0
        total_libras = registro['total_libras'] or 0
        porcentaje = (total_libras / total_cultivo_semana) * 100 if total_cultivo_semana else 0
        kilos = total_libras / 2.20462 if total_libras else 0

        resultado.append({
            'itemsapname': registro['itemsapname'],
            'categoria': registro['categoria'],
            'cultivo': registro['cultivo'],
            'semana': registro['semana'],
            'anio': registro['anio'],
            'libras': round(total_libras, 2),
            'cajas': round(total_cajas, 0),
            'porcentaje': round(porcentaje, 2),
            'kilos': round(kilos, 2),
        })

    registros_json = json.dumps(resultado, default=str)

    if not resultado:
        tabla_html = "<p class='text-danger'>No hay datos disponibles para mostrar.</p>"
    else:
        df = pd.DataFrame(resultado)

        metrica = request.GET.get('metrica', 'porcentaje')
        if metrica not in ['porcentaje', 'libras', 'kilos', 'cajas']:
            metrica = 'porcentaje'

        if metrica not in ['porcentaje', 'libras', 'kilos', 'cajas'] or metrica not in df.columns:
            tabla_html = f"<p class='text-warning'>Métrica '{metrica}' inválida o no disponible.</p>"
        else:
            tabla_pivote = df.pivot_table(
                index=['cultivo', 'categoria', 'itemsapname'],
                columns=['semana', 'anio'],
                values=metrica,
                aggfunc='sum'
            )
           
            tabla_pivote = tabla_pivote.fillna("")

            # Aplanar índices (convertir multiindex filas a columnas normales)
            tabla_pivote_reset = tabla_pivote.reset_index()

            # Aplanar multiindex de columnas, por ejemplo: (semana, anio) -> "semana_anio"
            tabla_pivote_reset.columns = [
                '_'.join(map(str, col)).strip() if isinstance(col, tuple) else col
                for col in tabla_pivote_reset.columns.values
            ]

            # Ahora sí, convertir a HTML
            tabla_html = tabla_pivote_reset.to_html(
                classes="table table-striped",
                index=False,
                na_rep="",
                table_id="tabla-pivote"
            )

    return render(request, 'plantaE/reportetecnicos/inventarioProd_reportesemanalprodterm_productor.html', {
        'tabla_html': tabla_html,
        'registros_json': registros_json,
    })

def recepciones_reporteAcum(request):
    today = timezone.now().date()

     # Obtener todos los registros para el usuario y la fecha
    registros = Recepciones.objects.filter(fecha=today)

   # Crear un DataFrame a partir de los registros, incluyendo todas las columnas
    df = pd.DataFrame(list(registros.values()),columns=['fecha','finca','cultivo','variedad','cajas','libras'])

    # Agrupar por 'variedad' y sumar las 'cajas'
    df_agrupado = df.groupby(['variedad','cultivo','finca'], as_index=False).agg(
        total_cajas=('cajas', 'sum'),
        cultivo=('cultivo', 'first'),  # Conservar el primer correo asociado
        fecha=('fecha', 'first'),
        variedad=('variedad', 'first'),
        finca=('finca', 'first'),
        total_libras=('libras', 'sum')    # Conservar la primera fecha asociada
    )

    # Convertir el DataFrame a una lista de diccionarios para pasarlo a la plantilla
    registros_finales = df_agrupado.to_dict(orient='records')

    # Agrupar por 'cultivo' y sumar las 'cajas'
    df_agrupado = df.groupby(['cultivo','finca'], as_index=False).agg(
        total_cajas=('cajas', 'sum'),
        cultivo=('cultivo', 'first'),  # Conservar el primer correo asociado
        fecha=('fecha', 'first'),
        variedad=('variedad', 'first'),
        finca=('finca', 'first'),
        total_libras=('libras', 'sum')    # Conservar la primera fecha asociada
    )

    registros_finales2 = df_agrupado.to_dict(orient='records')

    if request.method == 'POST':
        opcion2 = request.POST.get('opcion2')
         # Obtener todos los registros para el usuario y la fecha
        registros = Recepciones.objects.filter(fecha=opcion2)

    # Crear un DataFrame a partir de los registros, incluyendo todas las columnas
        df = pd.DataFrame(list(registros.values()),columns=['fecha','finca','cultivo','variedad','cajas','libras'])

        # Agrupar por 'variedad' y sumar las 'cajas'
        df_agrupado = df.groupby(['variedad','cultivo','finca'], as_index=False).agg(
            total_cajas=('cajas', 'sum'),
            cultivo=('cultivo', 'first'),  # Conservar el primer correo asociado
            fecha=('fecha', 'first'),
            variedad=('variedad', 'first'),
            finca=('finca', 'first'),
            total_libras=('libras', 'sum')    # Conservar la primera fecha asociada
        )

        # Convertir el DataFrame a una lista de diccionarios para pasarlo a la plantilla
        registros_finales = df_agrupado.to_dict(orient='records')

        # Agrupar por 'cultivo' y sumar las 'cajas'
        df_agrupado = df.groupby(['cultivo','finca'], as_index=False).agg(
            total_cajas=('cajas', 'sum'),
            cultivo=('cultivo', 'first'),  # Conservar el primer correo asociado
            fecha=('fecha', 'first'),
            variedad=('variedad', 'first'),
            finca=('finca', 'first'),
            total_libras=('libras', 'sum')    # Conservar la primera fecha asociada
        )

        registros_finales2 = df_agrupado.to_dict(orient='records')
        return JsonResponse({'datos': list(registros_finales),'opcion2':opcion2,'resumen':registros_finales2}, safe=False)

    return render(request, 'plantaE/reportetecnicos/recepciones_reporteAcum.html', {'registros': registros_finales, 'registros2': registros_finales2})

def recepciones_reporteAcumKgm2Orden(request):

    today = timezone.now().date()
    current_week = today.isocalendar()[1]  # Obtener el número de semana actual
    current_year = today.isocalendar()[0]  # Obtener el año actual

    # Obtener todos los registros
    registros = AcumFruta.objects.all()

    # Crear un DataFrame a partir de los registros
    df = pd.DataFrame(list(registros.values()), columns=['fecha', 'finca', 'orden','cultivo', 'libras'])
    df['fecha'] = pd.to_datetime(df['fecha'], errors='coerce')
    # Agregar columnas para el número de semana y el año
    df['semana'] = df['fecha'].dt.isocalendar().week
    df['año'] = df['fecha'].dt.isocalendar().year

    # Filtrar por la semana y el año actuales
    df_filtrado = df[(df['semana'] == current_week) & (df['año'] == current_year)]
    
    # Agrupar por 'variedad' y sumar las 'cajas'

    df_agrupado = df_filtrado.groupby(['orden', 'cultivo', 'finca'], as_index=False).agg(
        cultivo=('cultivo', 'first'),
        semana=('semana', 'first'),
        finca=('finca', 'first'),
        orden=('orden', 'first'),
        total_libras=('libras', 'sum')
    )
    df_agrupado['Kg'] = df_agrupado['total_libras']*0.453592
    areas = datosProduccion.objects.all()

    df_areas = pd.DataFrame(list(areas.values()), columns=['orden', 'area'])  # 'orden' y 'area'

    # Realizamos un merge para agregar las áreas correspondientes a cada 'orden'
    df_final = df_agrupado.merge(df_areas, on='orden', how='inner')
    df_final['kxm2'] = df_final['Kg']/df_final['area']
    # Convertir el DataFrame a una lista de diccionarios para pasarlo a la plantilla
    registros_finales = df_final.to_dict(orient='records')

    if request.method == 'POST':
        opcion2 = request.POST.get('opcion2')  # Esto es algo como "Semana 51 del 2024"
        opcion2 = opcion2.strip()  # Eliminar espacios adicionales o caracteres extraños
        # Obtener todos los registros para el usuario y la fecha
        registros = AcumFruta.objects.all()

        # Crear un DataFrame a partir de los registros
        df = pd.DataFrame(list(registros.values()), columns=['fecha', 'finca', 'orden', 'cultivo', 'libras'])
        df['fecha'] = pd.to_datetime(df['fecha'], errors='coerce')  # Convierte la fecha a datetime
        
        # Agregar columnas para el número de semana y el año
        df['semana'] = df['fecha'].dt.isocalendar().week  # Semana ISO
        df['año'] = df['fecha'].dt.isocalendar().year  # Año ISO
        df['semana-año'] = 'Semana ' + df['semana'].astype(str) + ' del ' + df['año'].astype(str)  # Texto para visualizar
        
        
        # Filtramos por las columnas 'semana-año'
        df_filtrado = df[df['semana-año'] == str(opcion2)]

        # Agrupar por 'orden', 'cultivo' y 'finca', sumando las libras
        df_agrupado = df_filtrado.groupby(['orden', 'cultivo', 'finca'], as_index=False).agg(
            cultivo=('cultivo', 'first'),
            semana=('semana', 'first'),
            finca=('finca', 'first'),
            orden=('orden', 'first'),
            total_libras=('libras', 'sum')
        )
        
        # Convertir las libras a kilogramos
        df_agrupado['Kg'] = df_agrupado['total_libras'] * 0.453592

        # Obtener las áreas
        areas = datosProduccion.objects.all()
        df_areas = pd.DataFrame(list(areas.values()), columns=['orden', 'area'])

        # Realizar un merge para agregar las áreas correspondientes a cada 'orden'
        df_final = df_agrupado.merge(df_areas, on='orden', how='inner')

        # Calcular kxm2 (kg por metro cuadrado)
        df_final['kxm2'] = df_final['Kg'] / df_final['area']

        # Convertir el DataFrame a una lista de diccionarios para pasarlo a la plantilla
        registros_finales = df_final.to_dict(orient='records')

        # Devolver los datos en formato JSON
        return JsonResponse({'datos': list(registros_finales), 'opcion2': opcion2}, safe=False)

    return render(request, 'plantaE/reportetecnicos/recepciones_reporteAcumKgm2Orden.html', {'registros': registros_finales})

def recepciones_reporteAcumKgm2Estruc(request):

    today = timezone.now().date()
    current_week = today.isocalendar()[1]  # Obtener el número de semana actual
    current_year = today.isocalendar()[0]  # Obtener el año actual

    # Obtener todos los registros
    registros = AcumFruta.objects.all()

    # Crear un DataFrame a partir de los registros
    df = pd.DataFrame(list(registros.values()), columns=['fecha', 'finca', 'orden','cultivo','estructura', 'libras'])
    df['fecha'] = pd.to_datetime(df['fecha'], errors='coerce')
    # Agregar columnas para el número de semana y el año
    df['semana'] = df['fecha'].dt.isocalendar().week
    df['año'] = df['fecha'].dt.isocalendar().year

    # Filtrar por la semana y el año actuales
    df_filtrado = df[(df['semana'] == current_week) & (df['año'] == current_year)]
    
    # Agrupar por 'variedad' y sumar las 'cajas'

    df_agrupado = df_filtrado.groupby(['orden', 'cultivo', 'finca','estructura'], as_index=False).agg(
        cultivo=('cultivo', 'first'),
        semana=('semana', 'first'),
        finca=('finca', 'first'),
        orden=('orden', 'first'),
        estructura=('estructura', 'first'),
        total_libras=('libras', 'sum')
    )
    df_agrupado['Kg'] = df_agrupado['total_libras']*0.453592
    areas = datosProduccion.objects.all()

    df_areas = pd.DataFrame(list(areas.values()), columns=['orden', 'area'])  # 'orden' y 'area'

    # Realizamos un merge para agregar las áreas correspondientes a cada 'orden'
    df_final = df_agrupado.merge(df_areas, on='orden', how='inner')
    df_final['kxm2'] = df_final['Kg']/df_final['area']
    # Convertir el DataFrame a una lista de diccionarios para pasarlo a la plantilla
    registros_finales = df_final.to_dict(orient='records')

    if request.method == 'POST':
        opcion2 = request.POST.get('opcion2')  # Esto es algo como "Semana 51 del 2024"
        opcion2 = opcion2.strip()  # Eliminar espacios adicionales o caracteres extraños
        # Obtener todos los registros para el usuario y la fecha
        registros = AcumFruta.objects.all()

        # Crear un DataFrame a partir de los registros
        df = pd.DataFrame(list(registros.values()), columns=['fecha', 'finca', 'orden', 'cultivo','estructura', 'libras'])
        df['fecha'] = pd.to_datetime(df['fecha'], errors='coerce')  # Convierte la fecha a datetime
        
        # Agregar columnas para el número de semana y el año
        df['semana'] = df['fecha'].dt.isocalendar().week  # Semana ISO
        df['año'] = df['fecha'].dt.isocalendar().year  # Año ISO
        df['semana-año'] = 'Semana ' + df['semana'].astype(str) + ' del ' + df['año'].astype(str)  # Texto para visualizar
        
        
        # Filtramos por las columnas 'semana-año'
        df_filtrado = df[df['semana-año'] == str(opcion2)]

        # Agrupar por 'orden', 'cultivo' y 'finca', sumando las libras
        df_agrupado = df_filtrado.groupby(['orden', 'cultivo', 'finca','estructura'], as_index=False).agg(
            cultivo=('cultivo', 'first'),
            semana=('semana', 'first'),
            finca=('finca', 'first'),
            orden=('orden', 'first'),
            estructura=('estructura', 'first'),
            total_libras=('libras', 'sum')
        )
        
        # Convertir las libras a kilogramos
        df_agrupado['Kg'] = df_agrupado['total_libras'] * 0.453592

        # Obtener las áreas
        areas = datosProduccion.objects.all()
        df_areas = pd.DataFrame(list(areas.values()), columns=['orden', 'area'])

        # Realizar un merge para agregar las áreas correspondientes a cada 'orden'
        df_final = df_agrupado.merge(df_areas, on='orden', how='inner')

        # Calcular kxm2 (kg por metro cuadrado)
        df_final['kxm2'] = df_final['Kg'] / df_final['area']

        # Convertir el DataFrame a una lista de diccionarios para pasarlo a la plantilla
        registros_finales = df_final.to_dict(orient='records')

        # Devolver los datos en formato JSON
        return JsonResponse({'datos': list(registros_finales), 'opcion2': opcion2}, safe=False)

    return render(request, 'plantaE/reportetecnicos/recepciones_reporteAcumKgm2Estruc.html', {'registros': registros_finales})

def recepciones_reporteAcumKgm2Variedad(request):

    today = timezone.now().date()
    current_week = today.isocalendar()[1]  # Obtener el número de semana actual
    current_year = today.isocalendar()[0]  # Obtener el año actual

    # Obtener todos los registros
    registros = AcumFruta.objects.all()

    # Crear un DataFrame a partir de los registros
    df = pd.DataFrame(list(registros.values()), columns=['fecha', 'finca', 'orden','cultivo','estructura','variedad', 'libras'])
    df['fecha'] = pd.to_datetime(df['fecha'], errors='coerce')
    # Agregar columnas para el número de semana y el año
    df['semana'] = df['fecha'].dt.isocalendar().week
    df['año'] = df['fecha'].dt.isocalendar().year

    # Filtrar por la semana y el año actuales
    df_filtrado = df[(df['semana'] == current_week) & (df['año'] == current_year)]
    
    # Agrupar por 'variedad' y sumar las 'cajas'

    df_agrupado = df_filtrado.groupby(['orden', 'cultivo', 'finca','variedad'], as_index=False).agg(
        cultivo=('cultivo', 'first'),
        semana=('semana', 'first'),
        finca=('finca', 'first'),
        orden=('orden', 'first'),
        variedad=('variedad', 'first'),
        total_libras=('libras', 'sum')
    )
    df_agrupado['Kg'] = df_agrupado['total_libras']*0.453592
    areas = datosProduccion.objects.all()

    df_areas = pd.DataFrame(list(areas.values()), columns=['orden', 'area'])  # 'orden' y 'area'

    # Realizamos un merge para agregar las áreas correspondientes a cada 'orden'
    df_final = df_agrupado.merge(df_areas, on='orden', how='inner')
    df_final['kxm2'] = df_final['Kg']/df_final['area']
    # Convertir el DataFrame a una lista de diccionarios para pasarlo a la plantilla
    registros_finales = df_final.to_dict(orient='records')

    if request.method == 'POST':
        opcion2 = request.POST.get('opcion2')  # Esto es algo como "Semana 51 del 2024"
        opcion2 = opcion2.strip()  # Eliminar espacios adicionales o caracteres extraños
        # Obtener todos los registros para el usuario y la fecha
        registros = AcumFruta.objects.all()

        # Crear un DataFrame a partir de los registros
        df = pd.DataFrame(list(registros.values()), columns=['fecha', 'finca', 'orden', 'cultivo','estructura','variedad', 'libras'])
        df['fecha'] = pd.to_datetime(df['fecha'], errors='coerce')  # Convierte la fecha a datetime
        
        # Agregar columnas para el número de semana y el año
        df['semana'] = df['fecha'].dt.isocalendar().week  # Semana ISO
        df['año'] = df['fecha'].dt.isocalendar().year  # Año ISO
        df['semana-año'] = 'Semana ' + df['semana'].astype(str) + ' del ' + df['año'].astype(str)  # Texto para visualizar
        
        
        # Filtramos por las columnas 'semana-año'
        df_filtrado = df[df['semana-año'] == str(opcion2)]

        # Agrupar por 'orden', 'cultivo' y 'finca', sumando las libras
        df_agrupado = df_filtrado.groupby(['orden', 'cultivo', 'finca','variedad'], as_index=False).agg(
            cultivo=('cultivo', 'first'),
            semana=('semana', 'first'),
            finca=('finca', 'first'),
            orden=('orden', 'first'),
            variedad=('variedad', 'first'),
            total_libras=('libras', 'sum')
        )
        
        # Convertir las libras a kilogramos
        df_agrupado['Kg'] = df_agrupado['total_libras'] * 0.453592

        # Obtener las áreas
        areas = datosProduccion.objects.all()
        df_areas = pd.DataFrame(list(areas.values()), columns=['orden', 'area'])

        # Realizar un merge para agregar las áreas correspondientes a cada 'orden'
        df_final = df_agrupado.merge(df_areas, on='orden', how='inner')

        # Calcular kxm2 (kg por metro cuadrado)
        df_final['kxm2'] = df_final['Kg'] / df_final['area']

        # Convertir el DataFrame a una lista de diccionarios para pasarlo a la plantilla
        registros_finales = df_final.to_dict(orient='records')

        # Devolver los datos en formato JSON
        return JsonResponse({'datos': list(registros_finales), 'opcion2': opcion2}, safe=False)

    return render(request, 'plantaE/reportetecnicos/recepciones_reporteAcumKgm2Variedad.html', {'registros': registros_finales})

def recepciones_reporteAcumSem(request):

    today = timezone.now().date()
    current_week = today.isocalendar()[1]  # Obtener el número de semana actual
    current_year = today.isocalendar()[0]  # Obtener el año actual

    # Obtener todos los registros
    registros = Recepciones.objects.all()

    # Crear un DataFrame a partir de los registros
    df = pd.DataFrame(list(registros.values()), columns=['fecha', 'finca', 'cultivo', 'variedad', 'cajas', 'libras'])
    df['fecha'] = pd.to_datetime(df['fecha'], errors='coerce')
    # Agregar columnas para el número de semana y el año
    df['semana'] = df['fecha'].dt.isocalendar().week
    df['año'] = df['fecha'].dt.isocalendar().year

    # Filtrar por la semana y el año actuales
    df_filtrado = df[(df['semana'] == current_week) & (df['año'] == current_year)]

    # Agrupar por 'variedad' y sumar las 'cajas'
    df_agrupado = df_filtrado.groupby(['variedad', 'cultivo', 'finca'], as_index=False).agg(
        total_cajas=('cajas', 'sum'),
        cultivo=('cultivo', 'first'),
        semana=('semana', 'first'),
        variedad=('variedad', 'first'),
        finca=('finca', 'first'),
        total_libras=('libras', 'sum')
    )

    # Convertir el DataFrame a una lista de diccionarios para pasarlo a la plantilla
    registros_finales = df_agrupado.to_dict(orient='records')

    # Agrupar por 'cultivo' y sumar las 'cajas'
    df_agrupado2 = df_filtrado.groupby(['cultivo', 'finca'], as_index=False).agg(
        total_cajas=('cajas', 'sum'),
        cultivo=('cultivo', 'first'),
        semana=('semana', 'first'),
        variedad=('variedad', 'first'),
        finca=('finca', 'first'),
        total_libras=('libras', 'sum')
    )

    registros_finales2 = df_agrupado2.to_dict(orient='records')

    return render(request, 'plantaE/reportetecnicos/recepciones_reporteAcumSem.html', {
        'registros': registros_finales,
        'registros2': registros_finales2
    })

def recepciones_reporteAcumSemPublic(request):
    today = timezone.now().date()
    current_week = today.isocalendar()[1]  # Obtener el número de semana actual
    current_year = today.isocalendar()[0]  # Obtener el año actual
    nombre_usuario = request.user.username
    # Obtener el objeto de usuario basado en el correo. Usamos .get() si esperamos un solo resultado

    try:
        finca_usuario = usuariosAppFruta.objects.get(correo=nombre_usuario)
        
    except usuariosAppFruta.DoesNotExist:
        # En caso de que no exista el usuario con ese correo, maneja el error apropiadamente.
        # Puedes lanzar una excepción o retornar un mensaje de error.
        finca_usuario = None

    # Verifica si se encontró el usuario
    if finca_usuario:
        # Obtener los registros de 'Recepciones' para la finca asociada al usuario
        registros = Recepciones.objects.filter(finca=finca_usuario.finca)

    # Crear un DataFrame a partir de los registros
    df = pd.DataFrame(list(registros.values()), columns=['fecha', 'finca', 'cultivo', 'variedad', 'cajas', 'libras'])
    df['fecha'] = pd.to_datetime(df['fecha'], errors='coerce')
    # Agregar columnas para el número de semana y el año
    df['semana'] = df['fecha'].dt.isocalendar().week
    df['año'] = df['fecha'].dt.isocalendar().year

    # Filtrar por la semana y el año actuales
    df_filtrado = df[(df['semana'] == current_week) & (df['año'] == current_year)]

    # Agrupar por 'variedad' y sumar las 'cajas'
    df_agrupado = df_filtrado.groupby(['variedad', 'cultivo', 'finca'], as_index=False).agg(
        total_cajas=('cajas', 'sum'),
        cultivo=('cultivo', 'first'),
        semana=('semana', 'first'),
        variedad=('variedad', 'first'),
        finca=('finca', 'first'),
        total_libras=('libras', 'sum')
    )

    # Convertir el DataFrame a una lista de diccionarios para pasarlo a la plantilla
    registros_finales = df_agrupado.to_dict(orient='records')

    # Agrupar por 'cultivo' y sumar las 'cajas'
    df_agrupado2 = df_filtrado.groupby(['cultivo', 'finca'], as_index=False).agg(
        total_cajas=('cajas', 'sum'),
        cultivo=('cultivo', 'first'),
        semana=('semana', 'first'),
        variedad=('variedad', 'first'),
        finca=('finca', 'first'),
        total_libras=('libras', 'sum')
    )

    registros_finales2 = df_agrupado2.to_dict(orient='records')

    return render(request, 'plantaE/reportetecnicos/recepciones_reporteAcumSemPublic.html', {
        'registros': registros_finales,
        'registros2': registros_finales2
    })

def recepciones_reportecurva(request):
    nombre_usuario = request.user.username
    #mensaje = request.POST.get('array')
    
    return render(request, 'plantaE/reportetecnicos/recepciones_reportegrafica.html', {'usuario': nombre_usuario})

def recepciones_reportecurva2(request):
    nombre_usuario = request.user.username
    #mensaje = request.POST.get('array')
    
    return render(request, 'plantaE/reportetecnicos/recepciones_reportegraficaPublic.html', {'usuario': nombre_usuario})

def reporte_tabla_pivote2(request):

    filtros_get = {
        'finca': request.GET.get('finca'),
        'orden': request.GET.get('orden'),
        'estructura': request.GET.get('estructura'),
        'variedad': request.GET.get('variedad'),
        'cultivo': request.GET.get('cultivo'),
    }

    nombre_usuario = request.user.username
    ordenes_abiertas = datosProduccion.objects.filter(status='Abierta').values_list('orden', flat=True)
    
    qs = AcumFruta.objects.filter(correo=nombre_usuario,orden__in=ordenes_abiertas).exclude(finca="CIP").exclude(libras__isnull=True).exclude(status="Anulado")

    for campo, valor in filtros_get.items():
        if valor:
            qs = qs.filter(**{campo: valor})

    data = qs.values('fecha', 'finca', 'orden','cultivo', 'estructura').annotate(
        total_libras=Sum('libras')
    )

    if data:
        df = pd.DataFrame(data)
        df['fecha'] = pd.to_datetime(df['fecha'])
        iso = df['fecha'].dt.isocalendar()
        df['semana'] = iso['year'].astype(str) + '-W' + iso['week'].astype(str).str.zfill(2)
        df['kg'] = df['total_libras'] / 2.20462  # convertir a kilogramos

        # Crear tabla pivote: kilos por semana
        pivot = pd.pivot_table(
            df,
            values='kg',
            index=['finca', 'orden','cultivo', 'estructura'],
            columns='semana',
            aggfunc='sum',
            fill_value=0
        ).round(2)

        # Obtener áreas por finca-orden-estructura
        areas_qs = detallesEstructuras.objects.values('finca', 'orden','cultivo', 'estructura').annotate(
            area_total=Sum('area')
        )
        df_areas = pd.DataFrame(list(areas_qs))

        # Unir pivot con áreas
        pivot = pivot.reset_index()
        df_merge = pd.merge(pivot, df_areas, on=['finca', 'orden', 'cultivo','estructura'], how='left')

        # Calcular rendimiento (kg/m²) por semana
        semanas = [col for col in df_merge.columns if col.startswith('20')]
        for semana in semanas:
            df_merge[semana] = df_merge[semana] / df_merge['area_total']
         
        df_merge[semanas] = df_merge[semanas].round(2)

        # Agregar columna total kilos por fila (sumar las semanas en kg)
        # Para total kilos, sumamos las semanas en el pivot original
        df_merge['total_kilos'] = pivot[semanas].sum(axis=1).round(2)

        # Agregar columna total kg/m² (sumar las semanas ya divididas por área)
        df_merge['total_kg_m2'] = df_merge[semanas].sum(axis=1).round(2)
        # Convertir resultado final a tabla HTML
        tabla_html = df_merge.to_html(
            classes='table table-striped table-bordered table-sm table-hover',
            index=False,
            table_id='tabla-pivote'
        )
    else:
        tabla_html = "<p>No hay datos para los filtros aplicados.</p>"

    # Filtros disponibles
    filtros_completos = [
        ('Finca', 'finca', AcumFruta.objects.filter(correo=nombre_usuario,orden__in=ordenes_abiertas).exclude(finca__isnull=True).exclude(finca='').values_list('finca', flat=True).distinct()),
        ('Orden', 'orden', AcumFruta.objects.filter(correo=nombre_usuario,orden__in=ordenes_abiertas).exclude(orden__isnull=True).exclude(orden='').values_list('orden', flat=True).distinct()),
        ('Variedad', 'variedad', AcumFruta.objects.filter(correo=nombre_usuario,orden__in=ordenes_abiertas).exclude(variedad__isnull=True).exclude(variedad='').values_list('variedad', flat=True).distinct()),
        ('Cultivo', 'cultivo', AcumFruta.objects.filter(correo=nombre_usuario,orden__in=ordenes_abiertas).exclude(cultivo__isnull=True).exclude(cultivo='').values_list('cultivo', flat=True).distinct()),
        ('Estructura', 'estructura', AcumFruta.objects.filter(correo=nombre_usuario,orden__in=ordenes_abiertas).exclude(estructura__isnull=True).exclude(estructura='').values_list('estructura', flat=True).distinct()),
    ]

    return render(request, 'plantaE/reportetecnicos/reporte_tabla_pivote2.html', {
        'tabla_html': tabla_html,
        'filtros_completos': filtros_completos,
        'request': request
    })

def reporte_tabla_pivote_produccionsem(request):

    nombre_usuario = request.user.username
    datos = usuariosAppFruta.objects.filter(correo=nombre_usuario).values('finca', 'encargado')

    # Definir fecha límite: 31 de octubre de 2025
    fecha_limite = datetime.date(2025, 9, 28)

    finca_usuario = datos[0]['finca']
    inventario_datos = AcumFruta.objects.filter(finca=finca_usuario,fecha__gt=fecha_limite).annotate(
        semana=ExtractWeek('fecha'),
        anio=ExtractIsoYear('fecha')
    ).values('orden', 'estructura', 'variedad', 'cultivo', 'semana', 'anio').annotate(
        total_libras=Sum('libras'),total_cajas=Sum('cajas')
    ).order_by('anio', 'semana', 'cultivo', 'estructura')

    resultado = []
    total_por_semana_cultivo = defaultdict(float)

    for registro in inventario_datos:
        clave = (registro['anio'], registro['semana'], registro['cultivo'])  # clave corregida
        total_libras = registro['total_libras'] or 0
        total_por_semana_cultivo[clave] += total_libras

    for registro in inventario_datos:
        clave_total = (registro['anio'], registro['semana'], registro['cultivo'])  # clave corregida
        total_cultivo_semana = total_por_semana_cultivo[clave_total]
        total_cajas = registro['total_cajas'] or 0
        total_libras = registro['total_libras'] or 0
        porcentaje = (total_libras / total_cultivo_semana) * 100 if total_cultivo_semana else 0
        kilos = total_libras / 2.20462 if total_libras else 0

        resultado.append({
            'orden': registro['orden'],
            'estructura': registro['estructura'],
            'cultivo': registro['cultivo'],
            'variedad': registro['variedad'],
            'semana': registro['semana'],
            'anio': registro['anio'],
            'libras': round(total_libras, 2),
            'cajas': round(total_cajas, 0),
            'porcentaje': round(porcentaje, 2),
            'kilos': round(kilos, 2),
        })

    registros_json = json.dumps(resultado, default=str)

    if not resultado:
        tabla_html = "<p class='text-danger'>No hay datos disponibles para mostrar.</p>"
    else:
        df = pd.DataFrame(resultado)

        metrica = request.GET.get('metrica', 'porcentaje')
        if metrica not in ['porcentaje', 'libras', 'kilos', 'cajas']:
            metrica = 'porcentaje'

        if metrica not in ['porcentaje', 'libras', 'kilos', 'cajas'] or metrica not in df.columns:
            tabla_html = f"<p class='text-warning'>Métrica '{metrica}' inválida o no disponible.</p>"
        else:
            tabla_pivote = df.pivot_table(
                index=['cultivo', 'estructura', 'variedad'],
                columns=['semana', 'anio'],
                values=metrica,
                aggfunc='sum'
            )
           
            tabla_pivote = tabla_pivote.fillna("")

            # Aplanar índices (convertir multiindex filas a columnas normales)
            tabla_pivote_reset = tabla_pivote.reset_index()

            # Aplanar multiindex de columnas, por ejemplo: (semana, anio) -> "semana_anio"
            tabla_pivote_reset.columns = [
                '_'.join(map(str, col)).strip() if isinstance(col, tuple) else col
                for col in tabla_pivote_reset.columns.values
            ]

            # Ahora sí, convertir a HTML
            tabla_html = tabla_pivote_reset.to_html(
                classes="table table-striped",
                index=False,
                na_rep="",
                table_id="tabla-pivote"
            )

    return render(request, 'plantaE/reportetecnicos/reporte_tabla_pivoteproduccionsem.html', {
        'tabla_html': tabla_html,
        'registros_json': registros_json,
    })

def aprovechamientos(request):
    hoy = timezone.now().date()
    semana_actual = hoy.isocalendar()[1]
    anio_actual = hoy.year

    # 1. Obtener fecha máxima en detallerecaux
    fecha_max = detallerecaux.objects.aggregate(max_fecha=Max('fechasalidafruta'))['max_fecha']
    if not fecha_max:
        fecha_max = hoy  # fallback si no hay registros

    # 2. Filtrar recepciones hasta esa fecha, semana actual
    recepciones = detallerec.objects.filter(
        fecha__lte=fecha_max
    ).annotate(
        semana=ExtractWeek('fecha'),
        anio=ExtractIsoYear('fecha')
    ).filter(
        semana=semana_actual,
        anio=anio_actual
    ).values('finca', 'cultivo').annotate(total_libras=Sum('libras')).order_by()

    # Crear dict con claves consistentes
    recepciones_dict = {
        formar_clave(r['finca'], r['cultivo']): r['total_libras'] for r in recepciones
    }

    # 3. Filtrar distribuciones por semana actual
    detalles = detallerecaux.objects.annotate(
        semana=ExtractWeek('fechasalidafruta'),
        anio=ExtractIsoYear('fechasalidafruta')
    ).filter(
        semana=semana_actual,
        anio=anio_actual
    )

    boleta_ids = detalles.values_list('boleta', flat=True)
    boletas = Boletas.objects.filter(boleta__in=boleta_ids)
    boletas_dict = {b.boleta: b for b in boletas}

    agrupados = defaultdict(lambda: {'aprovechamiento': 0, 'devolución': 0, 'mediano': 0, 'total': 0})
    detalle_debug = []

    for detalle in detalles:
        boleta = boletas_dict.get(detalle.boleta)
        if not boleta:
            continue
        if detalle.finca == "Productor":
            clave = formar_clave(detalle.llave, detalle.cultivo)
        else:
            clave = formar_clave(detalle.finca, detalle.cultivo)
        calidad = (boleta.calidad or '').strip().lower()
        libras = detalle.libras or 0

        if 'aprovechamiento' in calidad:
            agrupados[clave]['aprovechamiento'] += libras
        elif 'devolución' in calidad:
            agrupados[clave]['devolución'] += libras
        elif 'mediano' in calidad:
            agrupados[clave]['mediano'] += libras

        agrupados[clave]['total'] += libras

        # Solo para depuración: ejemplo con finca VALLE y cultivo GRAPE
        if clave == formar_clave("VALLE", "GRAPE"):
            detalle_debug.append({
                'boleta_id': detalle.boleta,
                'finca': detalle.finca,
                'cultivo': detalle.cultivo,
                'calidad': boleta.calidad,
                'libras': detalle.libras,
                'fecha_salida': detalle.fechasalidafruta,
            })

    # 4. Calcular resultados finales
    resultado = []
    for (finca, cultivo), datos in agrupados.items():
        total_distribuido = datos['total'] or 0
        recepcion_libras = recepciones_dict.get((finca, cultivo), 0)
        pendiente = recepcion_libras - total_distribuido
        if pendiente < 0:
            pendiente = 0
        porcentaje_pendiente = round(pendiente * 100 / recepcion_libras, 2) if recepcion_libras else 0

        resultado.append({
            'proveedor': finca,
            'cultivo': cultivo,
            'porcentaje_aprovechamiento': round(datos['aprovechamiento'] * 100 / total_distribuido, 2) if total_distribuido else 0,
            'porcentaje_devolucion': round(datos['devolución'] * 100 / total_distribuido, 2) if total_distribuido else 0,
            'porcentaje_mediano': round(datos['mediano'] * 100 / total_distribuido, 2) if total_distribuido else 0,
            'porcentaje_pendiente': porcentaje_pendiente,
            'libras_recibidas': round(recepcion_libras, 2),
            'libras_cargadas': round(total_distribuido, 2),
        })

    registros_json = json.dumps(resultado, default=str)
    detalle_debug_json = json.dumps(detalle_debug, default=str)

    return render(request, 'plantaE/reportetecnicos/inventarioProd_aprovechamientos.html', {
        'registros': resultado,
        'registros_json': registros_json,
        'detalle_debug': detalle_debug_json
    })

